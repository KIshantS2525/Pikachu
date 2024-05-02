import cv2
import os
import numpy as np
import openpyxl
from openpyxl import Workbook

# Set up parameters
size = 4
haar_file = "haarcascade_frontalface_default.xml"
datasets = "datasets"

# Load the face recognition model
model = cv2.face.LBPHFaceRecognizer_create()
model.read("trained_model.yml")  # Load a pre-trained model

# Load the cascade classifier for face detection
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# Load names and IDs
names = {}
with open("names.txt", "r") as f:
    for line in f:
        id, name = line.strip().split(",")
        names[int(id)] = name

# Create or load an Excel workbook
excel_file = "attendance.xlsx"
if os.path.exists(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active
else:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Date", "Time"] + list(names.values()))  # Header row

# Set up webcam
webcam = cv2.VideoCapture(0)
width, height = 130, 100

while True:
    ret, im = webcam.read()
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(30, 30))

    for (x, y, w, h) in faces:
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))
        prediction, confidence = model.predict(face_resize)

        if confidence < 100:
            recognized_name = names[prediction]
            cv2.putText(im, recognized_name, (x - 10, y - 10), cv2.FONT_HERSHEY_PLAIN, 1, (0, 255, 255), 2)
            
            # Record attendance in Excel
            row = [recognized_name] + [""] * len(names)
            row[worksheet.max_column - 1] = "X"  # Mark attendance with "X"
            worksheet.append(row)

        cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)
    
    cv2.imshow("Face Recognition", im)
    key = cv2.waitKey(10)
    if key == ord("q"):
        break

# Save the Excel workbook
workbook.save(excel_file)

webcam.release()
cv2.destroyAllWindows()
